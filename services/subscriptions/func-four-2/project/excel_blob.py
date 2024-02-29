from openpyxl import Workbook, load_workbook
from azure.storage.blob import BlobServiceClient
import io
import os
import sys

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from project.get_connection_string import get_connection_string_from_keyvault
import config.config_variables


def create_excel_blob():
    try:
        secret = config.config_variables.email_secret
        connection_string = get_connection_string_from_keyvault(secret)
        blob_name = "file_subscription.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Subscription Name"
        sheet["B1"] = "Subscription ID"
        sheet["C1"] = "Reason To Delete"
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        container_name = "excel"
        blob_service_client = BlobServiceClient.from_connection_string(
            connection_string
        )
        blob_client = blob_service_client.get_blob_client(
            container=container_name, blob=blob_name
        )
        blob_client.upload_blob(stream.getvalue(), overwrite=True)
    except Exception as ex:
        return str(ex)


def get_last_row():
    try:
        secret = config.config_variables.email_secret
        connection_string = get_connection_string_from_keyvault(secret)
        blob_service_client = BlobServiceClient.from_connection_string(
            connection_string
        )
        blob_name = "file_subscription.xlsx"
        container_name = "excel"
        container_client = blob_service_client.get_container_client(container_name)
        blob_client = container_client.get_blob_client(blob_name)
        blob_data = io.BytesIO(blob_client.download_blob().readall())
        workbook = load_workbook(blob_data)
        sheet = workbook.active
        last_row = sheet.max_row
        return last_row
    except Exception as ex:
        return str(ex)
